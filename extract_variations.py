import openpyxl
import json

wb = openpyxl.load_workbook('faire-products0.xlsx')
ws = wb['Products']

# Get column headers
headers = [cell.value for cell in ws[1]]

products = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    vals = {}
    for i, cell in enumerate(row):
        if i < len(headers) and headers[i]:
            vals[headers[i]] = cell.value
    
    name = vals.get('Product Name (English)')
    if not name:
        continue
    
    if name not in products:
        products[name] = {
            'name': name,
            'desc': vals.get('Description (English)'),
            'type': vals.get('Product Type'),
            'images': vals.get('Product Images'),
            'usd_retail': vals.get('USD Unit Retail Price'),
            'usd_wholesale': vals.get('USD Unit Wholesale Price'),
            'made_in': vals.get('Made In Country'),
            'variations': []
        }
    
    var_info = {}
    for opt_num in [1, 2, 3]:
        opt_name = vals.get(f'Option {opt_num} Name')
        opt_val = vals.get(f'Option {opt_num} Value')
        if opt_name and opt_val:
            var_info[opt_name] = opt_val
    
    if var_info:
        var_info['sku'] = vals.get('SKU')
        var_info['status'] = vals.get('Option Status')
        var_info['usd_retail'] = vals.get('USD Unit Retail Price')
        var_info['option_image'] = vals.get('Option Image')
        products[name]['variations'].append(var_info)

print(f"Total unique products: {len(products)}")
print()

# Show first 15 products with variation details
for i, (name, data) in enumerate(products.items()):
    if i >= 15:
        break
    print(f"--- Product {i+1}: {name} ---")
    print(f"  Type: {data['type']}")
    print(f"  USD Retail: {data['usd_retail']}")
    print(f"  Made In: {data['made_in']}")
    img_str = str(data['images']) if data['images'] else 'None'
    print(f"  Images: {img_str[:150]}")
    print(f"  Variations ({len(data['variations'])}):")
    for v in data['variations'][:8]:
        print(f"    {v}")
    if len(data['variations']) > 8:
        print(f"    ... and {len(data['variations'])-8} more")
    print()

# Collect all unique option names across products
all_option_names = set()
for p in products.values():
    for v in p['variations']:
        for k in v:
            if k not in ('sku', 'status', 'usd_retail', 'option_image'):
                all_option_names.add(k)

print(f"\n=== ALL UNIQUE VARIATION TYPES: {all_option_names} ===")

# Save full data to JSON
with open('faire_products_parsed.json', 'w', encoding='utf-8') as f:
    json.dump(list(products.values()), f, indent=2, ensure_ascii=False, default=str)
print(f"\nSaved {len(products)} products to faire_products_parsed.json")
