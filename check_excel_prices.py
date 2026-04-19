import openpyxl

wb = openpyxl.load_workbook('faire-products0.xlsx', read_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    # Print headers
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
        for i, h in enumerate(headers):
            print(f"  Col {i}: {h}")
    
    # Print first 3 data rows
    print("\n  First 3 rows of data:")
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True)):
        # Show only price-related cols
        price_cols = []
        for i, (h, v) in enumerate(zip(headers, row)):
            if h and ('price' in str(h).lower() or 'retail' in str(h).lower() or 'usd' in str(h).lower() or 'cost' in str(h).lower() or 'wholesale' in str(h).lower()):
                price_cols.append(f"{h}={v}")
        if price_cols:
            name = row[0] if row[0] else 'N/A'
            print(f"    Row {row_num+2}: {str(name)[:40]} | {' | '.join(price_cols)}")

wb.close()
